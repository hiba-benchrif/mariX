import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
import tempfile
from datetime import datetime

def format_excel_header(ws, headers):
    ws.append(headers)
    # Style de la première ligne
    fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        
    # Ajustement largeur des colonnes
    for col_num, column_title in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20

def generate_employee_report(user_id):
    from app.models import DossierImport, DossierExport
    
    wb = openpyxl.Workbook()
    
    # Onglet Imports
    ws_import = wb.active
    ws_import.title = "Mes Imports"
    headers = ["N° Dossier", "Client", "Fournisseur", "MBL", "POL", "POD", "Date Création", "Statut"]
    format_excel_header(ws_import, headers)
    
    imports = DossierImport.query.filter_by(created_by=user_id).all()
    for d in imports:
        ws_import.append([
            d.numero_dossier, d.client, d.fournisseur, d.mbl, d.pol, d.pod,
            d.created_at.strftime('%Y-%m-%d') if d.created_at else "",
            d.statut
        ])
        
    # Onglet Exports
    ws_export = wb.create_sheet(title="Mes Exports")
    headers = ["N° Dossier", "Booking", "Client", "Compagnie", "Date Chargement", "POD", "Date Création", "Statut"]
    format_excel_header(ws_export, headers)
    
    exports = DossierExport.query.filter_by(created_by=user_id).all()
    for d in exports:
        ws_export.append([
            d.numero_dossier, d.numero_booking, d.client, d.compagnie,
            d.date_chargement.strftime('%Y-%m-%d') if d.date_chargement else "",
            d.pod,
            d.created_at.strftime('%Y-%m-%d') if d.created_at else "",
            d.statut
        ])
        
    temp_dir = tempfile.gettempdir()
    filepath = os.path.join(temp_dir, f"Rapport_Employe_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
    wb.save(filepath)
    return filepath

def generate_weekly_report():
    from app.models import DossierImport, DossierExport, User
    from datetime import timedelta
    
    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday())
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Hebdomadaire"
    
    headers = ["Type", "N° Dossier", "Employé", "Client", "Date Création", "Statut"]
    format_excel_header(ws, headers)
    
    imports = DossierImport.query.filter(DossierImport.created_at >= start_of_week).all()
    for d in imports:
        user = User.query.get(d.created_by)
        employe_name = f"{user.prenom} {user.nom}" if user else "Inconnu"
        ws.append([
            "Import", d.numero_dossier, employe_name, d.client,
            d.created_at.strftime('%Y-%m-%d') if d.created_at else "",
            d.statut
        ])
        
    exports = DossierExport.query.filter(DossierExport.created_at >= start_of_week).all()
    for d in exports:
        user = User.query.get(d.created_by)
        employe_name = f"{user.prenom} {user.nom}" if user else "Inconnu"
        ws.append([
            "Export", d.numero_dossier, employe_name, d.client,
            d.created_at.strftime('%Y-%m-%d') if d.created_at else "",
            d.statut
        ])
        
    temp_dir = tempfile.gettempdir()
    filepath = os.path.join(temp_dir, f"Rapport_Hebdo_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
    wb.save(filepath)
    return filepath
